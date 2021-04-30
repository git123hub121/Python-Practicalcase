from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import openpyxl
import calendar

calendar.setfirstweekday(firstweekday=6)
wb = openpyxl.Workbook()
flex_text = []

def set_information(date, text):
    t = {}
    t['month'] = date.split('-')[1]
    t['day'] = date.split('-')[2]
    t['text'] = text
    flex_text.append(t)

set_information('2020-7-12', '考试')

for i in range(1, 8):
    # 添加工作表
    sheet = wb.create_sheet(index=0, title=str(i) + '月')
    # 获取具体日期时间
    count = 0
    for j in range(len(calendar.monthcalendar(2020, i))):
        for k in range(len(calendar.monthcalendar(2020, i)[j])):
            value = calendar.monthcalendar(2020, i)[j][k]
            # 将0值变为空值
            bd = Border(right=Side(color='00E6E6FA', style='thin'),
                        top=Side(color='00E6E6FA', style='thin'))
            right_bd = Border(right=Side(color='00E6E6FA', style='thin'))
            if value == 0:
                value = ''
                sheet.cell(row=j + 4 + 2*count, column=k + 1).value = value
                sheet.cell(row=j + 4 + 2*count, column=k + 1).border = bd
                sheet.cell(row=j + 5 + 2*count, column=k + 1).border = right_bd
                sheet.cell(row=j + 6 + 2*count, column=k + 1).border = right_bd
            else:
                sheet.cell(row=j + 4 + 2*count, column=k + 1).value = value
                sheet.cell(row=j + 4 + 2*count, column=k + 1).border = bd
                sheet.cell(row=j + 5 + 2*count, column=k + 1).border = right_bd
                sheet.cell(row=j + 6 + 2*count, column=k + 1).border = right_bd
                # 设置字体
                sheet.cell(row=j + 4 + 2*count, column=k + 1).font = Font(u'微软雅黑', color='8A2BE2' , size=14)
                # 单元格文字设置,右对齐,垂直居中
                
            for t in flex_text:
                if int(t['day']) == value and int(t['month']) == i:
                    print(t)
                    sheet.cell(row=j + 10 + 2*count, column=k + 1).value = t['text']
                    sheet.cell(row=j + 10 + 2*count, column=k + 1).font = Font(u'微软雅黑', color='8A2BE2', size=10)
                    align = Alignment(horizontal='right', vertical='center')
                    sheet.cell(row=j + 10 + 2*count, column=k + 1).alignment = align
        count += 1
    align = Alignment(horizontal='center', vertical='center')
    # 单元格填充色属性设置
    fill = PatternFill("solid", fgColor="FFFFFF")
    # 对单元格进行颜色填充
    for k1 in range(1, 30):
        for k2 in range(1, 30):
            sheet.cell(row=k1, column=k2).fill = fill
    # 添加星期几信息行
    days = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
    num = 0
    for k3 in range(1, 8):
        sheet.cell(row=3, column=k3).value = days[num]
        sheet.cell(row=3, column=k3).alignment = align
        sheet.cell(row=3, column=k3).font = Font(u'微软雅黑', color='8A2BE2', bold=True , size=12)
        # 设置列宽12
        c_char = get_column_letter(k3)
        sheet.column_dimensions[get_column_letter(k3)].width = 12
        # 行高27
        sheet.row_dimensions[k3].height = 27
        num += 1
    # 设置行高30
    for k4 in range(3, 19):
        sheet.row_dimensions[k4].height = 28
        # 合并单元格
    sheet.merge_cells('I1:P20')
    # 添加图片
    img = Image('2.png')

    sheet.add_image(img, 'I1')

    # 添加年份及月份
    sheet.cell(row=2, column=7).value = '2020'
    sheet.cell(row=2, column=7).font = Font(u'微软雅黑', size=30, color='8B008B')
    sheet.cell(row=2, column=1).value = str(i) + '月'
    sheet.cell(row=2, column=1).font = Font(u'微软雅黑', size=30, color='BA55D3')
    sheet.row_dimensions[2].height = 35


# 保存文档
wb.save('my_calendary.xlsx')